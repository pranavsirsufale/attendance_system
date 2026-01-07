import calendar, subprocess, os, sys, jwt, csv, re, traceback, logging
from django.shortcuts import render
import rest_framework
from django.conf import settings
from rest_framework.viewsets import ModelViewSet
from django.http import HttpResponse,StreamingHttpResponse
from django.utils import timezone
from rest_framework import viewsets, generics, status , serializers
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view , permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from app.permissions import IsAdmin
from django.db.models import Count, Q , F , ExpressionWrapper , FloatField , DecimalField, CharField , Value,Sum , Case, When, IntegerField
from django.contrib.auth.models import User
from datetime import datetime, timedelta
from .models import Session , Program, Attendance, Teacher, Student, Subject, Timetable, CalendarException , Section, ArchivalAttendance
from .serializers import (
    SessionSerializer, AttendanceSerializer , TimetableCreateSerializer, TeacherSerializer, StudentSerializer,
    TimetableSerializer, CalendarExceptionSerializer, ProgramSerializer , SubjectSerializer , SectionSerializer
    , AdminTeacherSerializer , AdminStudentSerializer , AdminTimetableSerializer,AttendanceStatsSerializer,
    StudentDetailSerializer, AttendanceSummarySerializer, SingleSessionTimetableSerializer, ArchivalAttendanceSerializer
)
from django.db.models.functions import Concat
from django.db import IntegrityError,transaction
from io import StringIO


logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s %(message)s', stream=sys.stdout)

class StandardPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

class StudentListView(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def get(self, request):
        try:
            students = Student.objects.all().order_by('roll_number')
            student_count = students.count()
            if student_count == 0:
                return Response({
                    "count": 0,
                    "next": null,
                    "previous": null,
                    "results": {
                        "students": [],
                        "total_count": 0
                    }
                }, status=200)

            paginator = self.pagination_class()
            paginated_students = paginator.paginate_queryset(students, request)
            serializer = StudentSerializer(paginated_students, many=True)
            return paginator.get_paginated_response({
                'students': serializer.data,
                'total_count': student_count
            })
        except Exception as e:
            logger.error(f"Error in StudentListView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class TeacherListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            teachers = Teacher.objects.all()
            serializer = TeacherSerializer(teachers, many=True)
            return Response(serializer.data, status=200)
        except Exception as e:
            logger.error(f"Error in TeacherListView: {str(e)}")
            return Response({"error": str(e)}, status=500)

class TeacherSubjectsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            teacher_id = request.query_params.get('teacher')
            if not teacher_id:
                return Response({"error": "Teacher ID is required"}, status=400)
            # Return all subjects assigned to the teacher via timetables
            from .models import Subject
            subject_qs = Subject.objects.filter(timetable__teacher_id=teacher_id).distinct().order_by('name')
            subject_list = list(subject_qs.values_list('name', flat=True))
            return Response({"subjects": subject_list}, status=200)
        except Exception as e:
            logger.error(f"Error in TeacherSubjectsView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class AdminCRUDViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_queryset(self):
        return self.queryset

class TeacherViewSet(AdminCRUDViewSet):
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer

class ProgramViewSet(AdminCRUDViewSet):
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer

class SubjectViewSet(AdminCRUDViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated]

class AdminAttendanceStatsView(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def get(self, request):
        try:
            period = request.query_params.get('period', 'semester')
            section_id = request.query_params.get('section')
            subject_id = request.query_params.get('subject')
            subject_name = request.query_params.get('subject_name')
            teacher_id = request.query_params.get('teacher')
            program_id = request.query_params.get('program')
            year = request.query_params.get('year')
            semester = request.query_params.get('semester')
            date = request.query_params.get('date')
            end_date = request.query_params.get('end_date')
            sessions = Session.objects.select_related('timetable__section__program', 'timetable__subject', 'timetable__teacher').filter(status='Completed')
            if section_id:
                sessions = sessions.filter(timetable__section_id=section_id)
            if subject_id:
                sessions = sessions.filter(timetable__subject_id=subject_id)
            if subject_name:
                sessions = sessions.filter(timetable__subject__name=subject_name)
            if teacher_id:
                sessions = sessions.filter(timetable__teacher_id=teacher_id)
            if program_id:
                sessions = sessions.filter(timetable__section__program_id=program_id)
            if year:
                sessions = sessions.filter(timetable__section__year=year)
            if semester:
                sessions = sessions.filter(timetable__subject__semester=semester)
            if period == 'daily' and date:
                start_date = date
                sessions = sessions.filter(date=date)
                if not end_date:
                    end_date = (datetime.strptime(date, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
            elif period == 'weekly':
                start_date = datetime.now().date() - timedelta(days=datetime.now().weekday())
                end_date = start_date + timedelta(days=7)
                sessions = sessions.filter(date__range=[start_date, end_date])
            elif period == 'monthly':
                start_date = datetime.now().date().replace(day=1)
                end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)
                sessions = sessions.filter(date__range=[start_date, end_date])
            elif period == 'semester' and semester:
                start_date = sessions.filter(timetable__subject__semester=semester).order_by('date').first().date if sessions.exists() else datetime.now().date()
                end_date = sessions.filter(timetable__subject__semester=semester).order_by('-date').first().date if sessions.exists() else datetime.now().date()
                sessions = sessions.filter(date__range=[start_date, end_date])
            elif period == "custom" and date and end_date:
                start_date = date
                sessions = sessions.filter(date__range=[start_date, end_date])
            else:
                start_date = sessions.order_by('date').first().date if sessions.exists() else datetime.now().date()
                end_date = sessions.order_by('-date').first().date if sessions.exists() else datetime.now().date()
            stats = (
                Attendance.objects.filter(session__in=sessions)
                .select_related('student', 'recorded_by')
                .values(
                    'student_id',
                    'student__first_name',
                    'student__last_name',
                    'student__roll_number',
                    'student__section__program__name',
                    'student__section__year',
                    'student__semester',
                    'session__timetable__subject__name',
                    'recorded_by__first_name',
                    'recorded_by__last_name'
                )
                .annotate(
                    name=Concat(F('student__first_name'), Value(' '), F('student__last_name'), output_field=CharField()),
                    roll_number=F('student__roll_number'),
                    program=F('student__section__program__name'),
                    year=F('student__section__year'),
                    semester=F('student__semester'),
                    subject_name=F('session__timetable__subject__name'),
                    recorded_by_name=Concat(F('recorded_by__first_name'), Value(' '), F('recorded_by__last_name'), output_field=CharField()),
                    total_sessions=Count('session'),
                    present=Count('session', filter=Q(status=True)),
                    absent=Count('session', filter=Q(status=False)),
                    attendance_percentage=ExpressionWrapper(
                        (F('present') * 100.0) / F('total_sessions'),
                        output_field=FloatField()
                    )
                )
                .values(
                    'student_id',
                    'name',
                    'roll_number',
                    'program',
                    'year',
                    'semester',
                    'subject_name',
                    'recorded_by_name',
                    'total_sessions',
                    'present',
                    'absent',
                    'attendance_percentage'
                )
                .order_by('roll_number')
            )

            viz_data = {
                'total_sessions': stats.aggregate(total=Count('total_sessions'))['total'] or 0,
                'total_present': stats.aggregate(total=Count('present'))['total'] or 0,
                'total_absent': stats.aggregate(total=Count('absent'))['total'] or 0,
                'by_subject': (
                    Attendance.objects.filter(session__in=sessions)
                    .values('session__timetable__subject__name')
                    .annotate(
                        present=Count('session', filter=Q(status=True)),
                        absent=Count('session', filter=Q(status=False))
                    )
                    .order_by('session__timetable__subject__name')
                ),
                'by_teacher': (
                    Attendance.objects.filter(session__in=sessions)
                    .values('session__timetable__teacher__first_name', 'session__timetable__teacher__last_name')
                    .annotate(
                        present=Count('session', filter=Q(status=True)),
                        absent=Count('session', filter=Q(status=False))
                    )
                    .order_by('session__timetable__teacher__first_name')
                ),
                'by_date': (
                    Attendance.objects.filter(session__in=sessions)
                    .values('session__date')
                    .annotate(
                        present=Count('session', filter=Q(status=True)),
                        absent=Count('session', filter=Q(status=False))
                    )
                    .order_by('session__date')
                ) if period in ['daily', 'weekly', 'monthly'] else [],
            }

            paginator = self.pagination_class()
            paginated_stats = paginator.paginate_queryset(stats, request)
            serializer = AttendanceStatsSerializer(paginated_stats, many=True)

            return paginator.get_paginated_response({
                'stats': serializer.data,
                'start_date': start_date,
                'end_date': end_date,
                'viz_data': viz_data,
            })
        except Exception as e:
            logger.error(f"Error in AdminAttendanceStatsView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class TeacherAttendanceStatsView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            teacher = Teacher.objects.get(user=request.user)
            period = request.query_params.get('period', 'semester')
            start_date_param = request.query_params.get('start_date', None)
            end_date_param = request.query_params.get('end_date', None)
            timetables = Timetable.objects.filter(teacher=teacher)
            session_ids = Session.objects.filter(timetable__in=timetables).values_list('id', flat=True)
            students = Student.objects.filter(section__in=timetables.values('section')).distinct()
            today = timezone.now().date()
            if start_date_param and end_date_param:
                start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
            elif period == 'weekly':
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
            elif period == 'monthly':
                start_date = today.replace(day=1)
                end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)
            else:  # semester
                if not timetables.exists():
                    start_date = today
                    end_date = today
                else:
                    start_date = min(t.semester_start_date for t in timetables)
                    end_date = max(t.semester_end_date for t in timetables)


            attendance_stats = (
    Attendance.objects.filter(session__id__in=session_ids, student__in=students)
    .filter(session__date__gte=start_date, session__date__lte=end_date)
    .values('student__id', 'student__first_name', 'student__last_name', 'student__roll_number')
    .annotate(
        total_sessions=Count('session'),
        present=Count('session', filter=Q(status=True)),
        absent=Count('session', filter=Q(status=False))
    )
    .order_by('student__roll_number')  # <-- Add this line
)
            stats = [
                {
                    'student_id': stat['student__id'],
                    'name': f"{stat['student__first_name']} {stat['student__last_name']}",
                    'roll_number': stat['student__roll_number'],
                    'total_sessions': stat['total_sessions'],
                    'present': stat['present'],
                    'absent': stat['absent'],
                    'attendance_percentage': round((stat['present'] / stat['total_sessions']) * 100, 2) if stat['total_sessions'] > 0 else 0
                }
                for stat in attendance_stats
            ]

            return Response({
                'period': period,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'stats': stats
            }, status=200)
        except Teacher.DoesNotExist:
            logger.error("Teacher not found for user: %s", request.user)
            return Response({"error": "Teacher not found"}, status=404)
        except Exception as e:
            logger.error("Error fetching attendance stats: %s", str(e))
            return Response({"error": str(e)}, status=500)

class SessionViewSet(viewsets.ModelViewSet):
    queryset = Session.objects.all()
    serializer_class = SessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Session.objects.filter(timetable__teacher__user=self.request.user)

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()

    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(recorded_by=Teacher.objects.get(user=self.request.user))

class TimetableViewSet(viewsets.ModelViewSet):
    queryset = Timetable.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return TimetableCreateSerializer
        return TimetableSerializer

    def get_queryset(self):
        return Timetable.objects.filter(teacher__user=self.request.user)

    def create(self, request, *args, **kwargs):
        try:
            mutable_data = request.data.copy()
            teacher = Teacher.objects.get(user=request.user)
            mutable_data['teacher'] = teacher.id
            serializer = self.get_serializer(data=mutable_data)
            serializer.is_valid(raise_exception=True)
            timetable_data = self.perform_create(serializer)
            return Response(timetable_data, status=status.HTTP_201_CREATED)
        except ValidationError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error("Unexpected error: %s\n%s", str(e), traceback.format_exc())
            return Response({"detail": f"Server error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def perform_create(self, serializer):
        validated_data = serializer.validated_data
        teacher = validated_data['teacher']
        daily_schedules = validated_data['daily_schedules']
        section = validated_data['section']
        semester_start_date = validated_data.get('semester_start_date')
        semester_end_date = validated_data.get('semester_end_date')

        if not daily_schedules:
            raise ValidationError("At least one daily schedule is required.")

        for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
            existing = Timetable.objects.filter(section=section, day_of_week=day, teacher=teacher).count()
            new_for_day = len([s for s in daily_schedules if s['day_of_week'] == day])
            if existing + new_for_day > 5:
                raise ValidationError(f"Cannot schedule more than 5 lectures on {day} for this section.")
            for schedule in [s for s in daily_schedules if s['day_of_week'] == day]:
                if Timetable.objects.filter(
                    teacher=teacher,
                    day_of_week=day,
                    start_time=schedule['start_time'],
                    semester_start_date=semester_start_date,
                    semester_end_date=semester_end_date
                ).exists():
                    raise ValidationError(f"Teacher already scheduled on {day} at {schedule['start_time']} for this semester.")

        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6}
        created_timetables = []
        for schedule in daily_schedules:
            timetable = Timetable.objects.create(
                section=section,
                teacher=teacher,
                subject=Subject.objects.get(id=schedule['subject']),
                day_of_week=schedule['day_of_week'],
                start_time=schedule['start_time'],
                semester_start_date=semester_start_date,
                semester_end_date=semester_end_date
            )
            created_timetables.append(timetable)
            start_date = semester_start_date
            end_date = semester_end_date
            current_date = start_date
            target_day = day_of_week_map[timetable.day_of_week]
            while current_date <= end_date:
                if current_date.weekday() == target_day:
                    Session.objects.get_or_create(
                        timetable=timetable,
                        date=current_date,
                        defaults={'status': 'Scheduled'}
                    )
                current_date += timedelta(days=1)

        timetable_serializer = TimetableSerializer(created_timetables, many=True)
        return timetable_serializer.data

class SingleSessionTimetableView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            serializer = SingleSessionTimetableSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data
            teacher = Teacher.objects.get(user=request.user)
            section = data['section']
            subject = data['subject']
            day_of_week = data['day_of_week']
            start_time = data['start_time']
            session_date = data['session_date']

            # Validate date matches day_of_week
            day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6}
            if session_date.weekday() != day_of_week_map[day_of_week]:
                return Response({"detail": f"Selected date {session_date} does not match {day_of_week}"}, status=status.HTTP_400_BAD_REQUEST)

            # Check for conflicts
            if Timetable.objects.filter(
                teacher=teacher,
                day_of_week=day_of_week,
                start_time=start_time,
                semester_start_date__lte=session_date,
                semester_end_date__gte=session_date
            ).exists():
                return Response({"detail": f"Teacher already scheduled on {day_of_week} at {start_time}"}, status=status.HTTP_400_BAD_REQUEST)

            # Create timetable with a single-day duration
            timetable = Timetable.objects.create(
                section=section,
                teacher=teacher,
                subject=subject,
                day_of_week=day_of_week,
                start_time=start_time,
                semester_start_date=session_date,
                semester_end_date=session_date
            )

            # Create single session
            session = Session.objects.create(
                timetable=timetable,
                date=session_date,
                status='Scheduled'
            )

            return Response(TimetableSerializer(timetable).data, status=status.HTTP_201_CREATED)
        except ValidationError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Teacher.DoesNotExist:
            return Response({"detail": "Teacher not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error("Unexpected error: %s\n%s", str(e), traceback.format_exc())
            return Response({"detail": f"Server error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

'''

class TimetableViewSet(viewsets.ModelViewSet):
    queryset = Timetable.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return TimetableCreateSerializer
        return TimetableSerializer

    def get_queryset(self):
        logger.debug("Fetching queryset for user: %s", self.request.user)
        return Timetable.objects.filter(teacher__user=self.request.user)

    def create(self, request, *args, **kwargs):
        logger.debug("POST request received at /api/timetables/")
        logger.debug("Request data: %s", request.data)
        try:
            mutable_data = request.data.copy()
            teacher = Teacher.objects.get(user=request.user)
            mutable_data['teacher'] = teacher.id
            serializer = self.get_serializer(data=mutable_data)
            serializer.is_valid(raise_exception=True)
            timetable_data = self.perform_create(serializer)
            logger.info("Timetable created successfully")
            return Response(timetable_data, status=status.HTTP_201_CREATED)
        except ValidationError as e:
            logger.error("Serializer validation failed: %s", str(e))
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error("Unexpected error: %s\n%s", str(e), traceback.format_exc())
            return Response({"detail": f"Server error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def perform_create(self, serializer):
        validated_data = serializer.validated_data
        teacher = validated_data['teacher']
        daily_schedules = validated_data['daily_schedules']
        section = validated_data['section']

        if not daily_schedules:
            raise ValidationError("At least one daily schedule is required.")

        for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
            existing = Timetable.objects.filter(section=section, day_of_week=day).count()
            new_for_day = len([s for s in daily_schedules if s['day_of_week'] == day])
            if existing + new_for_day > 5:
                raise ValidationError(f"Cannot schedule more than 5 lectures on {day} for this section.")
            for schedule in [s for s in daily_schedules if s['day_of_week'] == day]:
                if Timetable.objects.filter(
                    teacher=teacher,
                    day_of_week=day,
                    start_time=schedule['start_time'],
                    semester_start_date=validated_data['semester_start_date'],
                    semester_end_date=validated_data['semester_end_date']
                ).exists():
                    raise ValidationError(f"Teacher already scheduled on {day} at {schedule['start_time']} for this semester.")

        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5}
        created_timetables = []
        for schedule in daily_schedules:
            timetable = Timetable.objects.create(
                section=section,
                teacher=teacher,
                subject=Subject.objects.get(id=schedule['subject']),
                day_of_week=schedule['day_of_week'],
                start_time=schedule['start_time'],
                semester_start_date=validated_data['semester_start_date'],
                semester_end_date=validated_data['semester_end_date']
            )
            created_timetables.append(timetable)
            start_date = validated_data['semester_start_date']
            end_date = validated_data['semester_end_date']
            current_date = start_date
            target_day = day_of_week_map[timetable.day_of_week]
            while current_date <= end_date:
                if current_date.weekday() == target_day:
                    Session.objects.get_or_create(
                        timetable=timetable,
                        date=current_date,
                        defaults={'status': 'Scheduled'}
                    )
                current_date += timedelta(days=1)

        # Serialize the created timetables
        timetable_serializer = TimetableSerializer(created_timetables, many=True)
        return timetable_serializer.data

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = TimetableSerializer(instance, data=request.data, partial=True)  # Use TimetableSerializer for updates
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def perform_update(self, serializer):
        instance = serializer.instance
        Session.objects.filter(timetable=instance).delete()
        serializer.save()
        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5}
        start_date = instance.semester_start_date
        end_date = instance.semester_end_date
        current_date = start_date
        target_day = day_of_week_map[instance.day_of_week]
        while current_date <= end_date:
            if current_date.weekday() == target_day:
                Session.objects.get_or_create(
                    timetable=instance,
                    date=current_date,
                    defaults={'status': 'Scheduled'}
                )
            current_date += timedelta(days=1)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        Session.objects.filter(timetable=instance).delete()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


was working the first previous








class TimetableViewSet(viewsets.ModelViewSet):
    queryset = Timetable.objects.all()
    permission_classes = [IsAuthenticated]
    print(permission_classes)

    def get_serializer_class(self):
        if self.action in ['create']:
            return TimetableCreateSerializer
        return TimetableSerializer  # Includes semester_start_date, semester_end_date for updates

    def get_queryset(self):
        return Timetable.objects.filter(teacher__user=self.request.user)

    def perform_create(self, serializer):
        validated_data = serializer.validated_data
        teacher = validated_data['teacher']
        print(teacher)
        daily_schedules = validated_data['daily_schedules']
        section = validated_data['section']
        print(daily_schedules)
        print(section)
        if not daily_schedules:
            raise ValidationError("At least one daily schedule is required.")

        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5}
        created_timetables = []
        default_start = datetime.now().date()
        default_end = default_start + timedelta(days=180)  # 6 months default

        for schedule in daily_schedules:
            timetable = Timetable.objects.create(
                section=section,
                teacher=teacher,
                subject=Subject.objects.get(id=schedule['subject']),
                day_of_week=schedule['day_of_week'],
                start_time=schedule['start_time'],
                semester_start_date=default_start,
                semester_end_date=default_end
            )
            created_timetables.append(timetable)
            start_date = default_start
            end_date = default_end
            current_date = start_date
            target_day = day_of_week_map[timetable.day_of_week]
            while current_date <= end_date:
                if current_date.weekday() == target_day:
                    Session.objects.get_or_create(
                        timetable=timetable,
                        date=current_date,
                        defaults={'status': 'Scheduled'}
                    )
                current_date += timedelta(days=1)

        timetable_serializer = TimetableSerializer(created_timetables, many=True)
        return Response(timetable_serializer.data, status=status.HTTP_201_CREATED)

previous than above ( the second previous )

'''

class TeacherViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer
    permission_classes = [IsAuthenticated]

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def teacher_info(request):
    try:
        teacher = Teacher.objects.get(user=request.user)
        token = request.headers.get('Authorization', '').split(' ')[1]
        decoded = jwt.decode(token, options={"verify_signature": False})
        print(datetime.fromtimestamp(decoded['exp']).isoformat())
        return Response({
            'name': f"{teacher.first_name} {teacher.last_name}",
            'last_login': request.user.last_login.isoformat() if request.user.last_login else None,
            'token_expiry': datetime.fromtimestamp(decoded['exp']).isoformat() if 'exp' in decoded else 'Unknown',
            'is_admin': teacher.is_admin
        })
    except Teacher.DoesNotExist:
        return Response({'detail': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

class TeacherCalendarView(generics.ListAPIView):
    serializer_class = SessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        teacher = Teacher.objects.get(user=self.request.user)
        queryset = Session.objects.filter(timetable__teacher=teacher).order_by('date')
        section_id = self.request.query_params.get('section_id')
        if section_id:
            queryset = queryset.filter(timetable__section_id=section_id)
        return queryset

class MarkAttendanceView(generics.GenericAPIView):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, session_id):
        try:
            session = Session.objects.get(id=session_id)
            try:
                user_teacher = Teacher.objects.get(user=request.user)
            except Teacher.DoesNotExist:
                user_teacher = None

            # allow if the requesting user is the session teacher or an admin teacher
            if user_teacher is None:
                return Response({"error": "Teacher not found"}, status=status.HTTP_404_NOT_FOUND)
            if session.timetable.teacher != user_teacher and not user_teacher.is_admin:
                return Response({"error": "Not authorized to mark this session"}, status=status.HTTP_403_FORBIDDEN)

            students = Student.objects.filter(section=session.timetable.section)
            existing_attendance = Attendance.objects.filter(session=session).select_related('student')

            attendance_data = [
                {'student_id': a.student.id, 'status': 'Present' if a.status else 'Absent'}
                for a in existing_attendance
            ] if existing_attendance.exists() else []

            return Response({
                "session": SessionSerializer(session).data,
                "students": StudentSerializer(students, many=True).data,
                "attendance": attendance_data
            }, status=status.HTTP_200_OK)

        except Session.DoesNotExist:
            logger.error(f"Session {session_id} not found")
            return Response({"error": "Session not found"}, status=status.HTTP_404_NOT_FOUND)
        except Teacher.DoesNotExist:
            logger.error(f"Teacher not found for user: {request.user}")
            return Response({"error": "Teacher not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error in MarkAttendanceView GET: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request, session_id):
        try:
            session = Session.objects.get(id=session_id)
            try:
                user_teacher = Teacher.objects.get(user=request.user)
            except Teacher.DoesNotExist:
                user_teacher = None

            if user_teacher is None:
                return Response({"error": "Teacher not found"}, status=status.HTTP_404_NOT_FOUND)
            if session.timetable.teacher != user_teacher and not user_teacher.is_admin:
                return Response({"error": "Not authorized to mark this session"}, status=status.HTTP_403_FORBIDDEN)

            attendance_data = request.data.get('attendance', [])
            if not attendance_data:
                return Response({"error": "Attendance data required"}, status=status.HTTP_400_BAD_REQUEST)

            for entry in attendance_data:
                student = Student.objects.get(id=entry['student_id'])
                if student.section != session.timetable.section:
                    continue
                # Convert string status to boolean
                status_value = True if entry['status'] == 'Present' else False
                combined_datetime = datetime.combine(session.date, session.timetable.start_time)
                # If admin is marking on behalf, record the session's teacher as recorder
                recorded_by = user_teacher if not (user_teacher and user_teacher.is_admin) else session.timetable.teacher
                Attendance.objects.update_or_create(
                    student=student,
                    session=session,
                    defaults={
                        'status': status_value,
                        'recorded_by': recorded_by,
                        'timestamp': combined_datetime
                    })

            session.status = 'Completed'
            session.save()
            return Response({"message": "Attendance marked/updated successfully"}, status=status.HTTP_200_OK)
        except Session.DoesNotExist:
            return Response({"error": "Session not found"}, status=status.HTTP_404_NOT_FOUND)
        except Student.DoesNotExist:
            return Response({"error": "Invalid student ID"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error in MarkAttendanceView POST: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class HolidayListCreateView(generics.ListCreateAPIView):
    queryset = CalendarException.objects.all()
    serializer_class = CalendarExceptionSerializer
    permission_classes = [IsAuthenticated]
    def perform_create(self, serializer):
        holiday = serializer.save()
        Session.objects.filter(date=holiday.date).update(status='Cancelled')

class AttendanceStatsView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, roll_number):
        try:
            student = Student.objects.get(roll_number=roll_number)
            section = student.section
            sessions = Session.objects.filter(
                timetable__section=section, status='Completed'
            )
            total_sessions = sessions.count()
            attendance = Attendance.objects.filter(
                student=student, session__in=sessions
            ).aggregate(
                present=Count('id', filter=Q(status=True)),
                absent=Count('id', filter=Q(status=False))
            )
            monthly_stats = Attendance.objects.filter(
                student=student, session__in=sessions
            ).extra(
                select={'month': "EXTRACT(MONTH FROM session_date)"}
            ).values('month').annotate(
                present=Count('id', filter=Q(status=True)),
                absent=Count('id', filter=Q(status=False))
            )

            attendance_list = [
                'P' if a.status == 'Present' else 'A'
                for a in Attendance.objects.filter(student=student, session__in=sessions).order_by('session__date')
            ]

            return Response({
                'student': StudentSerializer(student).data,
                'total_sessions': total_sessions,
                'present': attendance['present'] or 0,
                'absent': attendance['absent'] or 0,
                'percentage': (attendance['present'] / total_sessions * 100) if total_sessions > 0 else 0,
                'monthly_stats': list(monthly_stats),
                'consolidated': ' '.join(attendance_list[:5])
            })
        except Student.DoesNotExist:
            return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

class ClassHourlyStatsView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, session_id):
        try:
            session = Session.objects.get(id=session_id)
            if session.timetable.teacher != Teacher.objects.get(user=request.user):
                return Response({"error": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)
            attendance = Attendance.objects.filter(session=session).aggregate(
                present=Count('id', filter=Q(status='Present')),
                absent=Count('id', filter=Q(status='Absent')))
            total_students = Student.objects.filter(section=session.timetable.section).count()
            return Response({
                'session': SessionSerializer(session).data,
                'present': attendance['present'] or 0,
                'absent': attendance['absent'] or 0,
                'total_students': total_students
            })
        except Session.DoesNotExist:
            return Response({"error": "Session not found"}, status=status.HTTP_404_NOT_FOUND)

class SectionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Section.objects.all()
    serializer_class = SectionSerializer
    permission_classes = [IsAuthenticated]

class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated]

class SubjectsForSectionView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        section_id = request.query_params.get('section_id')
        semester = request.query_params.get('semester')
        if not section_id or not semester:
            return Response({"error": "section_id and semester are required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            section_id_int = int(section_id)
            semester_int = int(semester)
            section = Section.objects.get(id=section_id_int)
        except ValueError as e:
            return Response({"error": "section_id and semester must be integers"}, status=status.HTTP_400_BAD_REQUEST)
        except Section.DoesNotExist:
            return Response({"error": f"Section with id={section_id} not found"}, status=status.HTTP_400_BAD_REQUEST)
        program_duration = section.program.duration_years * 2
        start_semester = (section.year - 1) * 2 + 1
        end_semester = min(section.year * 2, program_duration)
        if not (start_semester <= semester_int <= end_semester):
            logger.error(f"Semester {semester_int} invalid for section {section.id} (range: {start_semester}-{end_semester})")
            return Response({"error": f"Semester {semester_int} is invalid for {section.name} (valid: {start_semester}-{end_semester})"}, status=status.HTTP_400_BAD_REQUEST)
        subjects = Subject.objects.filter(semester=semester_int)
        if not subjects.exists():
            return Response({"message": "No subjects found", "data": []}, status=status.HTTP_200_OK)
        serializer = SubjectSerializer(subjects, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_sections(request):
    try:
        sections = Section.objects.all()
        serializer = SectionSerializer(sections, many=True)
        return Response(serializer.data)
    except Exception as e:
        logger.error(f"Error in get_sections: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_subjects_for_section(request):
    section_id = request.query_params.get('section_id')
    semester_start_date = request.query_params.get('semester_start_date')
    if not section_id or not semester_start_date:
        return Response({"error": "Section ID and semester start date are required"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        section = Section.objects.get(id=section_id)
        start_date = datetime.strptime(semester_start_date, '%Y-%m-%d').date()
        # Infer semester: Jan-Jun = odd (1, 3, 5), Jul-Dec = even (2, 4, 6)
        semester = (section.year * 2 - 1) if start_date.month <= 6 else (section.year * 2)
        # Filter subjects by semester and optionally by section-specific logic if Subject model has a section relation
        subjects = Subject.objects.filter(semester=semester)
        serializer = SubjectSerializer(subjects, many=True)
        return Response(serializer.data)
    except Section.DoesNotExist:
        return Response({"error": "Section not found"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in get_subjects_for_section: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_subjects(request):
    subjects = Subject.objects.all()
    serializer = SubjectSerializer(subjects, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_time_slots(request):
    time_slots = [slot[0] for slot in Timetable.LECTURE_SLOTS]
    return Response(time_slots)

class SectionSemesterWiseDataView(APIView):
    def get(self, request):
        data = {}
        sections = Section.objects.all()
        for section in sections:
            students = section.students.all()
            if not students.exists():
                continue
            section_data = {}
            for student in students:
                for subject in student.subjects.all():
                    sem = f"Semester {subject.semester}"
                    if sem not in section_data:
                        section_data[sem] = []
                    serialized = SubjectSerializer(subject).data
                    if serialized not in section_data[sem]:  # prevent duplicates
                        section_data[sem].append(serialized)
            data[str(section)] = section_data
        return Response(data)

#!                       ▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄
#!                       █    THE ADMIN VIEWS APPEAR HERE     █
#!                       ▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀▀

class AdminTeacherViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.all()
    serializer_class = AdminTeacherSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    def perform_create(self,serializer):
        validated_data = serializer.validated_data
        email = validated_data.get('email')
        first_name = validated_data.get('first_name')
        last_name = validated_data.get('last_name')
        phone = validated_data.get('phone','')
        password = validated_data.get('password')
        is_admin = validated_data.get('is_admin' , False)
        # Create a new User for the teacher
        # username = email
        # counter = 1
        # while User.objects.filter(username = username).exists():
        #     username = f'{base_username}{counter}'
        #     counter += 1
        try:
            new_user = User.objects.create_user(
                username = email , email = email , password = password, first_name = first_name,
                last_name = last_name, is_superuser = is_admin)
        except Exception as e:
            raise
        try:
            teacher = Teacher.objects.create( user = new_user, first_name = first_name, last_name = last_name,
                email = email, phone = phone, is_admin = is_admin)
            serializer.instance = teacher
        except Exception as e:
            logger.error(f"Failed to create teacher : {str(e)}")
            new_user.delete()
            raise

    def perform_update(self,serializer):
        teacher = self.get_object() # the teacher instance being updated
        user = teacher.user # the linked user instance
        validated_data = serializer.validated_data
        first_name = validated_data.get('first_name' , user.first_name)
        last_name = validated_data.get('last_name' , user.last_name)
        email = validated_data.get('email' , user.email)
        password = validated_data.get('password')  # May be None or Empty
        is_admin = validated_data.get('is_admin' , teacher.is_admin)
        try:
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.is_superuser = is_admin
            if password:            # Only update if provided and non-empty
                user.set_password(password)
            user.save()

        except Exception as e:
            logger.error(f"failed to update user : {str(e)}")
            raise

        # Update Teacher fields
        try:
            serializer.save(
                first_name = first_name ,
                last_name = last_name ,
                email = email ,
                phone = validated_data.get('phone' , teacher.phone),
                is_admin = is_admin
            )
        except Exception as e:
            logger.error(f"Failed to update teacher : {str(e)}")
            raise

    def destroy(self,request , *args,**kwargs):
        teacher = self.get_object()
        user = teacher.user
        try:
            teacher.delete()
            user.delete()
            return Response(status = status.HTTP_204_NO_CONTENT)
        except Exception as e:
            logger.error(f"Failed to delete teacher/user : {str(e)}")
            return Response({"error" : f"Failed to delete : {str(e)}"})

class SemesterForSectionViewTest(APIView):
    permission_classes = [IsAuthenticated,IsAdmin]
    def get(self,request):
        section = request.query_params.get('section_id')
        return Response({"data":'section has found good !!!!!!!!'})

class SemestersForSectionView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]
    def get(self, request):
        section_id = request.query_params.get('section_id')
        print('I got section Id here ',section_id)
        if not section_id :
            return Response({'error' : 'section_id is required '} , status = status.HTTP_400_BAD_REQUEST)
        try:
            section = Section.objects.get(id = int(section_id))
            program_duration = section.program.duration_years * 2  # e.g. 3 year = 6 sem
            start_semester = (section.year - 1 ) * 2 + 1
            end_semester = min( start_semester + 1 , program_duration)
            semesters = list(range(start_semester , end_semester + 1))
            return Response({'semesters' : semesters } , status = status.HTTP_200_OK)
        except Section.DoesNotExist:
            logger.error(f"Section id = {section_id} not found")
            return Response({'error' : f"Section with id = {section_id}"} , status = status.HTTP_404_NOT_FOUND)
        except ValueError:
            logger.error(f"Invalid section_id : {section_id}")
            return Response({'error' : 'Section_id must be an integer'} , status = status.HTTP_400_BAD_REQUEST)

class StudentViewSet(ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    pagination_class = StandardPagination
    def get_queryset(self):
        queryset = super().get_queryset()
        section_id = self.request.query_params.get('section_id')
        semester = self.request.query_params.get('semester')
        program_id = self.request.query_params.get('program_id')
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        if semester:
            queryset = queryset.filter(semester=semester)
        if program_id:
            queryset = queryset.filter(section__program_id=program_id)
        return queryset.select_related('section__program')

    def list(self, request, *args, **kwargs):
        if request.query_params.get('all') == 'true':
            # Return all students without pagination
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        return super().list(request, *args, **kwargs)

class StudentDetailView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]
    def get(self, request, pk):
        try:
            student = Student.objects.select_related('section__program').get(id=pk)
            serializer = StudentSerializer(student)
            attendance_query = Attendance.objects.filter(student=student).select_related('session__timetable__subject')
            date_param = request.query_params.get('date')
            week_param = request.query_params.get('week')
            month_param = request.query_params.get('month')
            semester_param = request.query_params.get('semester')
            start_date_param = request.query_params.get('start_date')
            end_date_param = request.query_params.get('end_date')
            if date_param:
                try:
                    date = datetime.strptime(date_param, '%Y-%m-%d').date()
                    attendance_query = attendance_query.filter(session__date=date)
                except ValueError:
                    return Response({'error': 'Invalid date format'}, status=400)
            elif week_param:
                try:
                    week_start = datetime.strptime(week_param, '%Y-%m-%d').date()
                    week_end = week_start + timedelta(days=6)
                    attendance_query = attendance_query.filter(session__date__range=[week_start, week_end])
                except ValueError:
                    return Response({'error': 'Invalid week format'}, status=400)
            elif month_param:
                try:
                    month_start = datetime.strptime(month_param, '%Y-%m').date()
                    month_end = month_start + relativedelta(months=1) - timedelta(days=1)
                    attendance_query = attendance_query.filter(session__date__range=[month_start, month_end])
                except ValueError:
                    return Response({'error': 'Invalid month format'}, status=400)
            elif semester_param:
                year = datetime.now().year
                semester = student.semester
                if semester % 2 == 1:  # Odd semester (Jul-Dec)
                    start_date = datetime(year, 7, 1).date()
                    end_date = datetime(year, 12, 31).date()
                else:  # Even semester (Jan-Jun)
                    start_date = datetime(year, 1, 1).date()
                    end_date = datetime(year, 6, 30).date()
                attendance_query = attendance_query.filter(session__date__range=[start_date, end_date])
            elif start_date_param and end_date_param:
                try:
                    start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
                    if start_date > end_date:
                        return Response({'error': 'Start date must be before end date'}, status=400)
                    attendance_query = attendance_query.filter(session__date__range=[start_date, end_date])
                except ValueError:
                    return Response({'error': 'Invalid date format'}, status=400)
            attendance_data = (
                attendance_query.values('session__timetable__subject__id', 'session__timetable__subject__name')
                .annotate(classes_attended=Count('id', filter=Q(status=True)), total_classes=Count('id')))
            attendance_serializer = AttendanceSummarySerializer(attendance_data, many=True)
            return Response({'student': serializer.data, 'attendance': attendance_serializer.data})
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=404)
        except Exception as e:
            logger.error(f"Error in StudentDetailView: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=500)

class AdminStudentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdmin]
    queryset = Student.objects.all()
    serializer_class = AdminStudentSerializer
    def get_queryset(self):
        queryset = super().get_queryset()
        section_id = self.request.query_params.get('section')
        semester = self.request.query_params.get('semester')
        if section_id:
            queryset = queryset.filter(section__id = section_id)
        if semester :
            queryset = queryset.filter(semester = semester)
        return queryset
    def perform_create(self,serializer):
        validated_data = serializer.validated_data
        section = validated_data['section']  # Now words with Primary related key
        semester = validated_data['semester']
        roll_number = validated_data.get('roll_number')
        if not roll_number:
            program_prefix = 'NG' if 'BALLB' in section.program.name else 'G'
            year_suffix = str(timezone.now().year)[-2:] # e.g. , "25" from 2025
            sequence = str(Student.objects.filter(section__program=section.program).count() + 1).zfill(3)
            roll_number = f"{program_prefix}{year_suffix}{sequence}"
        subjects = Subject.objects.filter(semester =semester)
        try:
            student = serializer.save(roll_number = roll_number)
            student.subjects.set(subjects)
        except Exception as e:
            logger.error(f"Failed to create student : {str(e)}")
            raise
    def perform_update(self, serializer):
        serializer.save()

class AdminProgramViewSet(viewsets.ModelViewSet):
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

class AdminSubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_queryset(self):
        queryset = super().get_queryset()
        semester = self.request.query_params.get('semester')
        if semester:
            try:
                queryset = queryset.filter(semester = int(semester))
            except ValueError:
                ### Handle invalid semester value gracefully
                queryset = queryset.none()
        return queryset

class AdminSectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.all()
    serializer_class = SectionSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_queryset(self):
        queryset = super().get_queryset()
        program_id = self.request.query_params.get('program')
        if program_id :
            try:
                program = Program.objects.get(id = program_id)
                max_year = program.duration_years
                queryset = queryset.filter(program__id = program_id , year__lte = max_year)
            except Program.DoesNotExist:
                logger.error(f"Program id = {program_id} not found ")
        return queryset

'''
previous one without update funcionlity

class AdminTimetableViewSet(viewsets.ModelViewSet):

    queryset = Timetable.objects.all()
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return TimetableCreateSerializer
        return TimetableSerializer

    def perform_create(self, serializer):
        validated_data = serializer.validated_data
        teacher = validated_data['teacher']
        daily_schedules = validated_data['daily_schedules']
        section = validated_data['section']
        semester_start_date = validated_data.get('semester_start_date')
        semester_end_date = validated_data.get('semester_end_date')

        if not daily_schedules:
            raise ValidationError("At least one daily schedule is required.")

        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5}
        created_timetables = []
        for schedule in daily_schedules:
            timetable = Timetable.objects.create(
                section=section,
                teacher=teacher,
                subject=Subject.objects.get(id=schedule['subject']),
                day_of_week=schedule['day_of_week'],
                start_time=schedule['start_time'],
                semester_start_date=semester_start_date,
                semester_end_date=semester_end_date
            )
            created_timetables.append(timetable)
            start_date = semester_start_date
            end_date = semester_end_date
            current_date = start_date
            target_day = day_of_week_map[timetable.day_of_week]
            while current_date <= end_date:
                if current_date.weekday() == target_day:
                    Session.objects.get_or_create(
                        timetable=timetable,
                        date=current_date,
                        defaults={'status': 'Scheduled'}
                    )
                current_date += timedelta(days=1)

        timetable_serializer = TimetableSerializer(created_timetables, many=True)
        return timetable_serializer.data



'''

class AdminTimetableViewSet(viewsets.ModelViewSet):
    queryset = Timetable.objects.all()
    permission_classes = [IsAuthenticated , IsAdmin]

    '''
    def get_serializer_class(self):
        if self.action in ['create' , 'update' , 'partial_update']:
            return TimetableCreateSerializer # Use for wirte operation
        return TimetableSerializer ### Use for read operation
    '''
    def get_serializer_class(self):
        if self.action == 'create' :
            return TimetableCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return AdminTimetableSerializer
        return TimetableSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        section_id = self.request.query_params.get('section_id')
        if section_id :
            queryset = queryset.filter(section_id = section_id)
        return queryset

    def perform_create(self,serializer):
        validated_data = serializer.validated_data
        teacher = validated_data['teacher']
        daily_schedules = validated_data['daily_schedules']
        section = validated_data['section']
        semester_start_date = validated_data.get('semester_start_date')
        semester_end_date = validated_data.get('semester_end_date')
        if not daily_schedules:
            raise ValidationError('At least one daily schedules is required ')
        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6}
        created_timetables = []
        for schedule in daily_schedules:
            timetable = Timetable.objects.create(
                section = section ,
                teacher = teacher ,
                subject = Subject.objects.get(id = schedule['subject']),
                day_of_week = schedule['day_of_week'] ,
                start_time = schedule['start_time'],
                semester_start_date = semester_start_date ,
                semester_end_date = semester_end_date
            )
            created_timetables.append(timetable)
            start_date = semester_start_date
            end_date = semester_end_date
            current_date = start_date
            target_day = day_of_week_map[timetable.day_of_week]
            while current_date <= end_date:
                if current_date.weekday() == target_day:
                    Session.objects.get_or_create(
                        timetable = timetable ,
                        date = current_date,
                        defaults = {'status' :'Scheduled'}
                    )
                current_date += timedelta(days = 1)

        timetable_serializer = TimetableSerializer(created_timetables , many = True)
        self.response_data = timetable_serializer.data

    def create(self,request,*args,**kwargs):
        serializer = self.get_serializer(data = request.data)
        serializer.is_valid(raise_exception = True)
        self.perform_create(serializer)
        return Response(self.response_data, status = status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        # Use all created timetables from context for response
        created_timetables = serializer.context.get('created_timetables', [serializer.instance])
        response_serializer = TimetableSerializer(created_timetables, many=True)
        return Response(response_serializer.data)

    '''
    def perform_update(self,serializer):
        instance = serializer.instance
        #? Regenerate sessions
        Session.objects.filter(timetable = instance).delete()
        serializer.save()
        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5}
        start_date = instance.semester_start_date
        end_date = instance.semester_end_date
        current_date = start_date
        target_day = day_of_week_map[instance.day_of_week]
        while current_date <= end_date:
            if current_date.weekday() == target_day :
                Session.objects.get_or_create(
                    timetable = instance ,
                    date = current_date ,
                    defaults = {'status' : 'Scheduled'}
                )
            current_date += timedelta(days = 1)

    '''

    def perform_update(self, serializer):
        instance = serializer.save()
        related_timetables = Timetable.objects.filter(
            section=instance.section,
            teacher=instance.teacher,
            semester_start_date=instance.semester_start_date,
            semester_end_date=instance.semester_end_date
        )
        Session.objects.filter(timetable__in=related_timetables).delete()
        day_of_week_map = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6}
        start_date = instance.semester_start_date
        end_date = instance.semester_end_date
        for timetable in related_timetables:
            current_date = start_date
            target_day = day_of_week_map[timetable.day_of_week]
            while current_date <= end_date:
                if current_date.weekday() == target_day:
                    Session.objects.get_or_create(
                        timetable=timetable,
                        date=current_date,
                        defaults={'status': 'Scheduled'}
                    )
                current_date += timedelta(days=1)

    def destroy(self,request , *args,**kwargs):
        instance = self.get_object()
        Session.objects.filter(timetable = instance).delete() ### Clear sessions
        self.perform_destroy(instance)
        return Response(status = status.HTTP_204_NO_CONTENT)

class SectionForProgramView(APIView):
    permission_classes = [IsAuthenticated,IsAdmin]

    def get(self,request):
        program_id = request.query_params.get('program_id')
        if not program_id:
            return Response({'error':"program_id is required" }, status = status.HTTP_400_BAD_REQUEST)
        try:
            sections = Section.objects.filter(program_id = program_id)
            print(sections)
            serializer = SectionSerializer(sections , many = True)
            return Response(serializer.data)
        except Program.DoesNotExist:
            return Response({"error":f"Program with ID {program_id} not found"}, status=status.HTTP_404_NOT_FOUND)
        except ValueError as e:
            return Response({"error":"the program Id must be an integer"},status = status.HTTP_400_BAD_REQUEST)

class SemestersForSectionView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        section_id = request.query_params.get('section_id')
        if not section_id:
            return Response({'error': 'section_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            section = Section.objects.get(id=int(section_id))
            program_duration = section.program.duration_years * 2
            start_semester = (section.year - 1) * 2 + 1
            end_semester = min(start_semester + 1, program_duration)
            semesters = list(range(start_semester, end_semester + 1))
            return Response({'semesters': semesters}, status=status.HTTP_200_OK)
        except Section.DoesNotExist:
            logger.error(f"Section id={section_id} not found")
            return Response({'error': f"Section with id={section_id} not found"}, status=status.HTTP_404_NOT_FOUND)
        except ValueError:
            logger.error(f"Invalid section_id: {section_id}")
            return Response({'error': 'section_id must be an integer'}, status=status.HTTP_400_BAD_REQUEST)

class AdminAttendanceOverview(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        period = request.query_params.get('period', 'semester')
        today = timezone.now().date()

        if period == 'weekly':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == 'monthly':
            start_date = today.replace(day=1)
            end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)
        else:  # semester
            timetables = Timetable.objects.all()
            start_date = min(t.semester_start_date for t in timetables) if timetables else today
            end_date = max(t.semester_end_date for t in timetables) if timetables else today

        stats = (
            Attendance.objects.filter(session__date__gte=start_date, session__date__lte=end_date)
            .values('session__timetable__section__name')
            .annotate(
                total_sessions=Count('session'),
                present=Count('session', filter=Q(status='Present')),
                absent=Count('session', filter=Q(status='Absent'))
            )
        )

        response = [
            {
                'section': stat['session__timetable__section__name'],
                'total_sessions': stat['total_sessions'],
                'present': stat['present'],
                'absent': stat['absent'],
                'attendance_percentage': round((stat['present'] / stat['total_sessions']) * 100, 2) if stat['total_sessions'] > 0 else 0
            }
            for stat in stats
        ]

        return Response({
            'period': period,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'stats': response
        })

class PassStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def post(self, request):
        try:
            students_data = request.data
            if not isinstance(students_data, list):
                return Response({"error": "Expected a list of students"}, status=400)

            errors = []
            updated_students = []

            with transaction.atomic():
                for student_data in students_data:
                    required_fields = ['id', 'semester', 'section_id']
                    missing_fields = [field for field in required_fields if field not in student_data or student_data[field] is None]
                    if missing_fields:
                        errors.append(f"ID {student_data.get('id', 'unknown')}: Missing {', '.join(missing_fields)}")
                        continue

                    try:
                        student = Student.objects.get(id=student_data['id'])
                    except Student.DoesNotExist:
                        errors.append(f"ID {student_data['id']}: Not found")
                        continue

                    try:
                        current_section = Section.objects.get(id=student.section_id)
                    except Section.DoesNotExist:
                        errors.append(f"ID {student_data['id']}: Current section invalid")
                        continue

                    try:
                        new_section = Section.objects.get(id=student_data['section_id'])
                    except Section.DoesNotExist:
                        errors.append(f"ID {student_data['id']}: Target section invalid")
                        continue

                    # Validate program consistency
                    if new_section.program_id != current_section.program_id:
                        errors.append(f"ID {student_data['id']}: Target section must belong to the same program")
                        continue

                    # Validate semester
                    max_semester = 6 if current_section.program_id == 2 else 10
                    new_semester = student_data['semester']
                    if new_semester < 1 or new_semester > max_semester:
                        errors.append(f"ID {student_data['id']}: Semester {new_semester} is out of range (1-{max_semester})")
                        continue

                    # Validate semester progression
                    if new_semester <= student.semester:
                        errors.append(f"ID {student_data['id']}: Target semester must be greater than current semester {student.semester}")
                        continue

                    student.semester = new_semester
                    student.section_id = student_data['section_id']
                    updated_students.append(student)

                if errors:
                    return Response({"error": "; ".join(errors)}, status=400)

                for student in updated_students:
                    student.save()

                return Response({"updated": len(updated_students)}, status=200)

        except Exception as e:
            logger.error(f"Error in PassStudentsView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class RemoveStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def post(self, request):
        try:
            students_data = request.data
            if not isinstance(students_data, list):
                return Response({"error": "Expected a list of student IDs"}, status=400)

            errors = []
            deleted_count = 0

            with transaction.atomic():
                for student_data in students_data:
                    if 'id' not in student_data or student_data['id'] is None:
                        errors.append(f"Missing ID in student data")
                        continue

                    try:
                        student = Student.objects.get(id=student_data['id'])
                        student.delete()
                        deleted_count += 1
                    except Student.DoesNotExist:
                        errors.append(f"ID {student_data['id']}: Not found")
                        continue

                if errors:
                    return Response({"error": "; ".join(errors)}, status=400)

                return Response({"removed": deleted_count}, status=200)

        except Exception as e:
            logger.error(f"Error in RemoveStudentsView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class AdminAttendanceExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        try:
            period = request.query_params.get('period', 'semester')
            section_id = request.query_params.get('section')
            subject_id = request.query_params.get('subject')
            subject_name = request.query_params.get('subject_name')
            teacher_id = request.query_params.get('teacher')
            program_id = request.query_params.get('program')
            year = request.query_params.get('year')
            semester = request.query_params.get('semester')
            date = request.query_params.get('date')
            end_date = request.query_params.get('end_date')
            start_date = request.query_params.get('start_date')
            format = request.query_params.get('format', 'csv')
            if not teacher_id or not subject_name:
                return Response({"error": "Teacher ID and subject name are required"}, status=400)
            try:
                teacher = Teacher.objects.get(id=teacher_id)
            except Teacher.DoesNotExist:
                logger.error(f"Teacher id={teacher_id} not found")
                return Response({"error": f"Teacher with id={teacher_id} not found"}, status=404)

            sessions = Session.objects.select_related('timetable__section__program', 'timetable__subject', 'timetable__teacher').filter(status='Completed')
            if section_id:
                sessions = sessions.filter(timetable__section_id=section_id)
            if subject_id:
                sessions = sessions.filter(timetable__subject_id=subject_id)
            if subject_name:
                sessions = sessions.filter(timetable__subject__name=subject_name)
            if teacher_id:
                sessions = sessions.filter(timetable__teacher_id=teacher_id)
            if program_id:
                sessions = sessions.filter(timetable__section__program_id=program_id)
            if year:
                sessions = sessions.filter(timetable__section__year=year)
            if semester:
                sessions = sessions.filter(timetable__subject__semester=semester)

            if start_date and end_date:
                try:
                    sessions = sessions.filter(date__range=[start_date,end_date])
                except ValueError:
                    return Response({"error":"Invalid date range"},status = 400)
            if period == 'daily' and date:
                try:
                    sessions = sessions.filter(date=date)
                except ValueError:
                    logger.error(f"Invalid date format: {date}")
                    return Response({"error": "Invalid date format"}, status=400)
                if not end_date:
                    end_date = (datetime.strptime(date, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
            elif period == 'weekly':
                start_date = datetime.now().date() - timedelta(days=datetime.now().weekday())
                end_date = start_date + timedelta(days=7)
                sessions = sessions.filter(date__range=[start_date, end_date])
            elif period == 'monthly':
                start_date = datetime.now().date().replace(day=1)
                end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)
                sessions = sessions.filter(date__range=[start_date, end_date])
            elif period == 'semester' and semester:
                try:
                    start_date = sessions.filter(timetable__subject__semester=semester).order_by('date').first().date if sessions.exists() else datetime.now().date()
                    end_date = sessions.filter(timetable__subject__semester=semester).order_by('-date').first().date if sessions.exists() else datetime.now().date()
                    sessions = sessions.filter(date__range=[start_date, end_date])
                except ValueError:
                    logger.error(f"Invalid semester: {semester}")
                    return Response({"error": "Invalid semester"}, status=400)
            else:
                start_date = sessions.order_by('date').first().date if sessions.exists() else datetime.now().date()
                end_date = sessions.order_by('-date').first().date if sessions.exists() else datetime.now().date()
            stats = (
                Attendance.objects.filter(session__in=sessions)
                .select_related('student', 'recorded_by')
                .values(
                    'student_id',
                    'student__first_name',
                    'student__last_name',
                    'student__roll_number',
                    'student__section__program__name',
                    'student__section__year',
                    'student__semester',
                    'session__timetable__subject__name',
                    'recorded_by__first_name',
                    'recorded_by__last_name'
                )
                .annotate(
                    name=Concat(F('student__first_name'), Value(' '), F('student__last_name'), output_field=CharField()),
                    roll_number=F('student__roll_number'),
                    program=F('student__section__program__name'),
                    year=F('student__section__year'),
                    semester=F('student__semester'),
                    subject_name=F('session__timetable__subject__name'),
                    recorded_by_name=Concat(F('recorded_by__first_name'), Value(' '), F('recorded_by__last_name'), output_field=CharField()),
                    total_sessions=Count('session'),
                    present=Count('session', filter=Q(status=True)),
                    absent=Count('session', filter=Q(status=False)),
                    attendance_percentage=ExpressionWrapper(
                        (F('present') * 100.0) / F('total_sessions'),
                        output_field=FloatField()
                    )
                )
                .values(
                    'student_id',
                    'name',
                    'roll_number',
                    'program',
                    'year',
                    'semester',
                    'subject_name',
                    'recorded_by_name',
                    'total_sessions',
                    'present',
                    'absent',
                    'attendance_percentage'
                )
                .order_by('roll_number')
            )
            if format == 'csv':
                def stream_csv():
                    buffer = StringIO()
                    writer = csv.writer(buffer)
                    writer.writerow([
                        'Student Name', 'Student ID', 'Roll Number', 'Program', 'Year', 'Semester',
                        'Subject', 'Total Sessions', 'Present', 'Absent', 'Attendance %', 'Recorded By'
                    ])

                    for stat in stats.iterator():
                        writer.writerow([
                            stat['name'],
                            stat['student_id'],
                            stat['roll_number'],
                            stat['program'],
                            stat['year'],
                            stat['semester'],
                            stat['subject_name'],
                            stat['total_sessions'],
                            stat['present'],
                            stat['absent'],
                            f"{stat['attendance_percentage']:.2f}%",
                            stat['recorded_by_name']
                        ])
                        buffer.seek(0)
                        yield buffer.read()
                        buffer.truncate(0)
                        buffer.seek(0)
                response = StreamingHttpResponse(stream_csv(), content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="attendance_report_{subject_name}_{teacher_id}.csv"'
                return response
            else:
                return Response({"error": "Only CSV format is supported"}, status=400)
        except Exception as e:
            logger.error(f"Error in AdminAttendanceExportView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class AdminHolidayManagement(generics.ListCreateAPIView):
    queryset = CalendarException.objects.all()
    serializer_class = CalendarExceptionSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def perform_create(self, serializer):
        holiday = serializer.save()
        Session.objects.filter(date=holiday.date).update(status='Cancelled')

class BulkStudentUploadView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def post(self, request):
        try:
            students_data = request.data
            if not isinstance(students_data, list):
                return Response({"error": "Expected a list of students"}, status=400)

            # Validate data
            errors = []
            student_objects = []
            existing_ids = set(Student.objects.values_list('id', flat=True))

            for i, student in enumerate(students_data):
                # Required fields
                required_fields = ['id', 'first_name', 'roll_number', 'section_id', 'semester']
                missing_fields = [field for field in required_fields if field not in student or student[field] is None]
                if missing_fields:
                    errors.append(f"Row {i+1}: Missing fields: {', '.join(missing_fields)}")
                    continue

                # Validate ID
                if not isinstance(student['id'], int):
                    errors.append(f"Row {i+1}: Invalid ID (must be integer)")
                    continue
                if student['id'] in existing_ids:
                    errors.append(f"Row {i+1}: Duplicate ID {student['id']}")
                    continue

                # Validate section
                try:
                    section = Section.objects.get(id=student['section_id'])
                except Section.DoesNotExist:
                    errors.append(f"Row {i+1}: Invalid section_id")
                    continue

                # Validate semester (1-10 based on section year)
                year = section.year
                valid_semesters = [year * 2 - 1, year * 2]
                if student['semester'] not in valid_semesters:
                    errors.append(f"Row {i+1}: Invalid semester for year {year}")
                    continue

                # Validate email format if provided
                if student.get('email') and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', student['email']):
                    errors.append(f"Row {i+1}: Invalid email format")
                    continue

                student_objects.append(Student(
                    id=student['id'],
                    first_name=student['first_name'],
                    last_name=student.get('last_name', ''),
                    roll_number=student['roll_number'],
                    email=student.get('email', ''),
                    phone=student.get('phone', ''),
                    section_id=student['section_id'],
                    semester=student['semester']
                ))
            if errors:
                return Response({"error": "; ".join(errors)}, status=400)
            # Bulk create students
            try:
                Student.objects.bulk_create(student_objects)
                return Response({"added": len(student_objects)}, status=201)
            except IntegrityError as e:
                logger.error(f"IntegrityError during bulk create: {str(e)}")
                return Response({"error": "Database error, possible duplicate data"}, status=400)

        except Exception as e:
            logger.error(f"Error in BulkStudentUploadView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class TestExportView(APIView):
    permission_classes = [IsAuthenticated,IsAdmin]
    def get(self,request):
        return Response({"message":"Test export view works"})

class AdminSessionViewSet(viewsets.ModelViewSet):
    queryset = Session.objects.all()
    serializer_class = SessionSerializer
    permission_classes = [IsAuthenticated , IsAdmin]
    pagination_class = StandardPagination

    def get_queryset(self):
        queryset = super().get_queryset()
        section_id = self.request.query_params.get('section_id')
        teacher_id = self.request.query_params.get('teacher_id')
        semester = self.request.query_params.get('semester')

        if section_id :
            queryset = queryset.filter( timetable__section_id = section_id)
        if teacher_id :
            queryset = queryset.filter( timetable__teacher_id = teacher_id)
        if semester :
            queryset = queryset.filter( timetable__subject__semester = semester)

        return queryset.order_by('date')

    def get(self, request):
        try:
            date = request.query_params.get('date')
            section_id = request.query_params.get('section_id')
            # support explicit teacher_id (admin) and subject_name filters
            teacher_param = request.query_params.get('teacher') or request.query_params.get('teacher_id')
            subject_name = request.query_params.get('subject_name')
            if not date:
                return Response({"error": "Date parameter is required (format: yyyy-MM-dd)"}, status=400)
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return Response({"error": "Invalid date format. Use yyyy-MM-dd"}, status=400)
            # build base queryset depending on teacher filters
            if teacher_param:
                try:
                    teacher_obj = Teacher.objects.get(id=int(teacher_param))
                    sessions = Session.objects.filter(date=date_obj, timetable__teacher=teacher_obj)
                except Teacher.DoesNotExist:
                    return Response({"error": "Teacher not found"}, status=404)
            else:
                # fallback to current user's teacher if exists
                try:
                    teacher = Teacher.objects.get(user=request.user)
                    sessions = Session.objects.filter(date=date_obj, timetable__teacher=teacher)
                except Teacher.DoesNotExist:
                    sessions = Session.objects.filter(date=date_obj)
            if subject_name:
                sessions = sessions.filter(timetable__subject__name=subject_name)
            if section_id:
                sessions = sessions.filter(timetable__section_id=section_id)
            serializer = SessionSerializer(sessions, many=True)
            return Response(serializer.data, status=200)

        except Exception as e:
            logger.error(f"Error in SessionsByDateView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class ScheduledDatesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            month = request.query_params.get('month')
            section_id = request.query_params.get('section_id')
            # optional filters
            teacher_param = request.query_params.get('teacher') or request.query_params.get('teacher_id')
            subject_name = request.query_params.get('subject_name')
            if not month:
                return Response({"error": "Month parameter is required (format: yyyy-MM)"}, status=400)
            try:
                year, month = map(int, month.split('-'))
                start_date = datetime(year, month, 1).date()
                _, last_day = calendar.monthrange(year, month)
                end_date = datetime(year, month, last_day).date()
            except ValueError:
                return Response({"error": "Invalid month format. Use yyyy-MM"}, status=400)
            # determine teacher filter
            if teacher_param:
                try:
                    teacher_obj = Teacher.objects.get(id=int(teacher_param))
                    sessions = Session.objects.filter(date__range=[start_date, end_date], timetable__teacher=teacher_obj)
                except Teacher.DoesNotExist:
                    return Response({"error": "Teacher not found"}, status=404)
            else:
                try:
                    teacher = Teacher.objects.get(user=request.user)
                    sessions = Session.objects.filter(date__range=[start_date, end_date], timetable__teacher=teacher)
                except Teacher.DoesNotExist:
                    # admin or non-teacher: return all sessions in range
                    sessions = Session.objects.filter(date__range=[start_date, end_date])
            if section_id:
                sessions = sessions.filter(timetable__section_id=section_id)
            if subject_name:
                sessions = sessions.filter(timetable__subject__name=subject_name)
            dates = sessions.values('date').distinct().values_list('date', flat=True)
            date_list = [date.strftime('%Y-%m-%d') for date in dates]
            return Response({"dates": date_list}, status=200)

        except Exception as e:
            logger.error(f"Error in ScheduledDatesView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class SessionsByDateView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            date = request.query_params.get('date')
            section_id = request.query_params.get('section_id')
            # optionally allow specifying a teacher id (for admin browsing)
            teacher_param = request.query_params.get('teacher') or request.query_params.get('teacher_id')
            subject_name = request.query_params.get('subject_name')
            if not date:
                return Response({"error": "Date parameter is required (format: yyyy-MM-dd)"}, status=400)
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return Response({"error": "Invalid date format. Use yyyy-MM-dd"}, status=400)
            # determine teacher filter: either explicit param or current user
            if teacher_param:
                try:
                    teacher_obj = Teacher.objects.get(id=int(teacher_param))
                except Teacher.DoesNotExist:
                    return Response({"error": "Teacher not found"}, status=404)
                sessions = Session.objects.filter(date=date_obj, timetable__teacher=teacher_obj)
            else:
                try:
                    teacher = Teacher.objects.get(user=request.user)
                    sessions = Session.objects.filter(date=date_obj, timetable__teacher=teacher)
                except Teacher.DoesNotExist:
                    # if current user is not a teacher and no teacher param provided, return all sessions for date
                    sessions = Session.objects.filter(date=date_obj)
            if subject_name:
                sessions = sessions.filter(timetable__subject__name=subject_name)
            if section_id:
                sessions = sessions.filter(timetable__section_id=section_id)
            serializer = SessionSerializer(sessions, many=True)
            return Response(serializer.data, status=200)

        except Exception as e:
            logger.error(f"Error in SessionsByDateView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class StudentAttendanceView(APIView):
    permission_classes = [IsAuthenticated]
    # pagination_class = StandardPagination
    def get(self, request):
        try:
            program_id = request.query_params.get('program_id')
            section_id = request.query_params.get('section_id')
            semester = request.query_params.get('semester')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            if not program_id or not section_id or not semester:
                return Response({"error": "program_id, section_id, and semester are required"}, status=400)
            try:
                program_id = int(program_id)
                section_id = int(section_id)
                semester = int(semester)
            except ValueError:
                return Response({"error": "program_id, section_id, and semester must be integers"}, status=400)
            try:
                section = Section.objects.get(id=section_id, program_id=program_id)
            except Section.DoesNotExist:
                logger.error(f"Section {section_id} or Program {program_id} not found")
                return Response({"error": f"Section {section_id} or Program {program_id} not found"}, status=404)
            program_duration = section.program.duration_years * 2
            start_semester = (section.year - 1) * 2 + 1
            end_semester = min(start_semester + 1, program_duration)
            if not (start_semester <= semester <= end_semester):
                logger.error(f"Semester {semester} invalid for section {section_id}")
                return Response({"error": f"Semester {semester} is invalid (valid: {start_semester}-{end_semester})"}, status=400)
            subjects = Subject.objects.filter(semester=semester, timetable__section_id=section_id, timetable__section__program_id=program_id).distinct()
            if not subjects.exists():
                return Response({"count": 0, "subjects": [], "students": []}, status=200)
            students = Student.objects.filter(section_id=section_id, semester=semester)
            if not students.exists():
                return Response({
                    "count": 0,
                    "subjects": SubjectSerializer(subjects, many=True).data,
                    "students": []
                }, status=200)

            # Apply date range filter if provided
            sessions = Session.objects.filter(
                timetable__section_id=section_id,
                timetable__subject__semester=semester,
                timetable__section__program_id=program_id,
                status='Completed'
            )
            if start_date and end_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if start_date > end_date:
                        return Response({"error": "start_date must be before end_date"}, status=400)
                    sessions = sessions.filter(date__range=[start_date, end_date])
                except ValueError:
                    logger.error(f"Invalid date format: start_date={start_date}, end_date={end_date}")
                    return Response({"error": "Invalid date format (use YYYY-MM-DD)"}, status=400)

            # Fetch attendance data
            attendance_data = (
                Attendance.objects.filter(session__in=sessions, student__in=students)
                .values(
                    'student_id',
                    'student__first_name',
                    'student__last_name',
                    'student__roll_number',
                    'session__timetable__subject__id',
                    'session__timetable__subject__name'
                )
                .annotate(
                    name=Concat(F('student__first_name'), Value(' '), F('student__last_name'), output_field=CharField()),
                    classes_attended=Count('id', filter=Q(status=True)),
                    total_classes=Count('id')
                )
            )

            # Organize attendance by student
            student_attendance = {}
            for entry in attendance_data:
                student_id = entry['student_id']
                if student_id not in student_attendance:
                    student_attendance[student_id] = {
                        'id': student_id,
                        'roll_number': entry['student__roll_number'],
                        'name': entry['name'],
                        'attendance': []
                    }
                student_attendance[student_id]['attendance'].append({
                    'subject_id': entry['session__timetable__subject__id'],
                    'subject_name': entry['session__timetable__subject__name'],
                    'classes_attended': entry['classes_attended'],
                    'total_classes': entry['total_classes']
                })

            # Prepare student data
            student_data = list(student_attendance.values())
            # paginator = self.pagination_class()
            # paginated_students = paginator.paginate_queryset(student_data, request)
            # print(paginated_students)

            # return paginator.get_paginated_response({
            return Response({
                'count' : len(student_data),
                "results": {
                'subjects': SubjectSerializer(subjects, many=True).data,
                'students': student_data,
                'start_date': start_date.isoformat() if start_date else None,
                'end_date': end_date.isoformat() if end_date else None
            }})

        except Exception as e:
            logger.error(f"Error in StudentAttendanceView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

class MySQLBackupView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request, *args, **kwargs):
        db_config = settings.DATABASES.get('default')

        if not db_config:
            return Response(
                {"error": "Default database configuration not found in settings.py."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        MYSQL_HOST = db_config.get('HOST', 'localhost')
        MYSQL_USER = db_config.get('USER')
        MYSQL_PASSWORD = db_config.get('PASSWORD')
        MYSQL_DATABASE = db_config.get('NAME')
        if not all([MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE]):
            return Response(
                {"error": "Missing essential MySQL database credentials (USER, PASSWORD, NAME) in settings.py."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        BACKUP_DIR = getattr(settings, 'MYSQL_BACKUP_DIR', os.path.join(settings.BASE_DIR, 'mysql_backups'))
        # 1. Ensure the backup directory exists
        try:
            os.makedirs(BACKUP_DIR, exist_ok=True)
        except OSError as e:
            print(f"Error creating backup directory {BACKUP_DIR}: {e}")
            return Response(
                {"error": f"Failed to create backup directory on server: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file_name = f"{MYSQL_DATABASE}_backup_{timestamp}.sql"
        backup_path = os.path.join(BACKUP_DIR, backup_file_name)

        try:
            command = [
                'mysqldump',
                f'--host={MYSQL_HOST}',
                f'--user={MYSQL_USER}',
                f'--password={MYSQL_PASSWORD}',
                '--single-transaction',
                MYSQL_DATABASE
            ]

            process = subprocess.run(
                command,
                capture_output=True,
                text=True,
                check=True
            )

            with open(backup_path, 'w', encoding='utf-8') as f:
                f.write(process.stdout)

            print(f"MySQL backup successful: {backup_path}")
            return Response(
                {
                    "message": f"Database '{MYSQL_DATABASE}' backed up successfully.",
                    "filename": backup_file_name,
                    "path": backup_path
                },
                status=status.HTTP_200_OK
            )

        except subprocess.CalledProcessError as e:
            error_message = f"MySQL backup command failed. Stderr: {e.stderr.strip()}"
            print(f"Error during mysqldump execution: {error_message}")
            return Response(
                {"error": error_message},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except FileNotFoundError:
            error_message = "'mysqldump' command not found. Please ensure it's installed and in your server's PATH."
            print(f"Error: {error_message}")
            return Response(
                {"error": error_message},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            error_message = f"An unexpected server error occurred during backup: {str(e)}"
            print(f"An unexpected error occurred: {error_message}")
            return Response(
                {"error": error_message},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ArchivalAttendanceView(APIView):
    """
    Admin-only view for managing archival attendance records.
    GET: Retrieve archived attendance with filtering
    POST: Create archival snapshot from existing attendance records
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    
    def get(self, request):
        """
        Retrieve archived attendance with optional filters:
        - student_roll_number: Filter by student
        - subject_name: Filter by subject
        - section_name: Filter by section
        - semester: Filter by semester
        - start_date: Filter by session date (from)
        - end_date: Filter by session date (to)
        - archived_after: Filter by archive date (from)
        """
        try:
            archives = ArchivalAttendance.objects.all()
            
            # Apply filters
            student_roll = request.query_params.get('student_roll_number')
            if student_roll:
                archives = archives.filter(student_roll_number__icontains=student_roll)
            
            subject = request.query_params.get('subject_name')
            if subject:
                archives = archives.filter(subject_name__icontains=subject)
            
            section = request.query_params.get('section_name')
            if section:
                archives = archives.filter(section_name__icontains=section)
            
            semester = request.query_params.get('semester')
            if semester:
                archives = archives.filter(semester=semester)
            
            start_date = request.query_params.get('start_date')
            if start_date:
                archives = archives.filter(session_date__gte=start_date)
            
            end_date = request.query_params.get('end_date')
            if end_date:
                archives = archives.filter(session_date__lte=end_date)
            
            archived_after = request.query_params.get('archived_after')
            if archived_after:
                archives = archives.filter(archived_at__gte=archived_after)
            
            # Pagination
            paginator = StandardPagination()
            paginated_archives = paginator.paginate_queryset(archives, request)
            serializer = ArchivalAttendanceSerializer(paginated_archives, many=True)
            
            return paginator.get_paginated_response(serializer.data)
        
        except Exception as e:
            return Response(
                {"error": f"Failed to retrieve archival attendance: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def post(self, request):
        """
        Create archival snapshot from existing attendance records.
        Required fields:
        - filters: Dict with filters (program_id, semester, start_date, end_date)
        - archive_note: Optional note about this archive
        
        This will snapshot all matching attendance records into the archive.
        """
        try:
            teacher = Teacher.objects.get(user=request.user)
            
            filters = request.data.get('filters', {})
            archive_note = request.data.get('archive_note', '')
            
            # Build attendance query
            attendance_query = Attendance.objects.select_related(
                'student', 'student__section', 'student__section__program',
                'session', 'session__timetable', 'session__timetable__subject',
                'recorded_by'
            ).all()
            
            # Apply filters
            program_id = filters.get('program_id')
            if program_id:
                attendance_query = attendance_query.filter(student__section__program_id=program_id)
            
            semester = filters.get('semester')
            if semester:
                attendance_query = attendance_query.filter(student__semester=semester)
            
            start_date = filters.get('start_date')
            if start_date:
                attendance_query = attendance_query.filter(session__date__gte=start_date)
            
            end_date = filters.get('end_date')
            if end_date:
                attendance_query = attendance_query.filter(session__date__lte=end_date)
            
            # Create archival records
            archival_records = []
            for attendance in attendance_query:
                recorded_by_name = "Unknown"
                if attendance.recorded_by:
                    recorded_by_name = f"{attendance.recorded_by.first_name} {attendance.recorded_by.last_name}"
                
                archival_record = ArchivalAttendance(
                    student_roll_number=attendance.student.roll_number,
                    student_name=f"{attendance.student.first_name} {attendance.student.last_name}",
                    section_name=str(attendance.student.section),
                    subject_name=attendance.session.timetable.subject.name,
                    session_date=attendance.session.date,
                    semester=attendance.student.semester,
                    status=attendance.status,
                    original_timestamp=attendance.timestamp,
                    original_recorded_by=recorded_by_name,
                    archived_by=teacher,
                    archive_note=archive_note
                )
                archival_records.append(archival_record)
            
            if not archival_records:
                return Response(
                    {"error": "No attendance records found matching the specified filters"},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Bulk create archival records
            ArchivalAttendance.objects.bulk_create(archival_records)
            
            return Response({
                "message": f"Successfully archived {len(archival_records)} attendance records",
                "count": len(archival_records),
                "archived_by": f"{teacher.first_name} {teacher.last_name}",
                "archived_at": timezone.now()
            }, status=status.HTTP_201_CREATED)
        
        except Teacher.DoesNotExist:
            return Response(
                {"error": "Teacher profile not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to create archival snapshot: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ArchivalAttendanceStatsView(APIView):
    """
    Admin-only view for getting statistics about archived attendance.
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    
    def get(self, request):
        """
        Get summary statistics about archived attendance records.
        """
        try:
            total_archived = ArchivalAttendance.objects.count()
            
            # Stats by semester
            semester_stats = ArchivalAttendance.objects.values('semester').annotate(
                count=Count('id'),
                present_count=Sum(Case(When(status=True, then=1), default=0, output_field=IntegerField())),
                absent_count=Sum(Case(When(status=False, then=1), default=0, output_field=IntegerField()))
            ).order_by('semester')
            
            # Recent archives
            recent_archives = ArchivalAttendance.objects.values(
                'archived_at', 'archived_by__first_name', 'archived_by__last_name'
            ).annotate(
                count=Count('id')
            ).order_by('-archived_at')[:10]
            
            # Format recent archives
            recent_formatted = []
            for archive in recent_archives:
                recent_formatted.append({
                    'archived_at': archive['archived_at'],
                    'archived_by': f"{archive['archived_by__first_name']} {archive['archived_by__last_name']}" if archive['archived_by__first_name'] else "Unknown",
                    'count': archive['count']
                })
            
            return Response({
                "total_archived_records": total_archived,
                "semester_statistics": list(semester_stats),
                "recent_archives": recent_formatted
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {"error": f"Failed to retrieve archival statistics: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ArchivalSemesterSummaryView(APIView):
    """
    Admin-only: Return semester-wise subject summary from archival records.
    Query params: semester (required), section_name (optional)
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        try:
            semester = request.query_params.get('semester')
            if not semester:
                return Response({"error": "semester query parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

            archives = ArchivalAttendance.objects.filter(semester=semester)
            section = request.query_params.get('section_name')
            if section:
                archives = archives.filter(section_name__icontains=section)

            stats = (
                archives.values('subject_name')
                .annotate(
                    present_count=Sum(Case(When(status=True, then=1), default=0, output_field=IntegerField())),
                    absent_count=Sum(Case(When(status=False, then=1), default=0, output_field=IntegerField())),
                    total_records=Count('id')
                )
                .order_by('subject_name')
            )

            # compute percentage safely
            result = []
            for s in stats:
                total = s['total_records'] or 0
                present = s['present_count'] or 0
                percent = (present * 100.0 / total) if total > 0 else 0.0
                result.append({
                    'subject_name': s['subject_name'],
                    'present': present,
                    'absent': s['absent_count'] or 0,
                    'total_records': total,
                    'present_percent': round(percent, 2)
                })

            return Response({'semester': int(semester), 'subjects': result}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error("Error in ArchivalSemesterSummaryView: %s", str(e), exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ArchivalSubjectStudentsView(APIView):
    """
    Admin-only: For a given semester and subject, return per-student totals.
    Query params: semester (required), subject_name (required), section_name (optional)
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        try:
            semester = request.query_params.get('semester')
            subject = request.query_params.get('subject_name')
            if not semester or not subject:
                return Response({"error": "semester and subject_name are required"}, status=status.HTTP_400_BAD_REQUEST)

            archives = ArchivalAttendance.objects.filter(semester=semester, subject_name__icontains=subject)
            section = request.query_params.get('section_name')
            if section:
                archives = archives.filter(section_name__icontains=section)

            students = (
                archives.values('student_roll_number', 'student_name')
                .annotate(
                    present=Sum(Case(When(status=True, then=1), default=0, output_field=IntegerField())),
                    absent=Sum(Case(When(status=False, then=1), default=0, output_field=IntegerField())),
                    total=Count('id')
                )
                .order_by('student_roll_number')
            )

            result = []
            for st in students:
                total = st['total'] or 0
                present = st['present'] or 0
                percent = (present * 100.0 / total) if total > 0 else 0.0
                result.append({
                    'student_roll_number': st['student_roll_number'],
                    'student_name': st['student_name'],
                    'present': present,
                    'absent': st['absent'] or 0,
                    'total_records': total,
                    'present_percent': round(percent, 2)
                })

            return Response({'semester': int(semester), 'subject_name': subject, 'students': result}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error("Error in ArchivalSubjectStudentsView: %s", str(e), exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ArchivalStudentDetailView(APIView):
    """
    Admin-only: For a given student (roll) and subject+semester return detailed attendance list and summary.
    Query params: semester (required), subject_name (required), student_roll_number (required)
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        try:
            semester = request.query_params.get('semester')
            subject = request.query_params.get('subject_name')
            roll = request.query_params.get('student_roll_number')
            if not semester or not subject or not roll:
                return Response({"error": "semester, subject_name and student_roll_number are required"}, status=status.HTTP_400_BAD_REQUEST)

            archives = ArchivalAttendance.objects.filter(semester=semester, subject_name__icontains=subject, student_roll_number__icontains=roll).order_by('session_date')

            total = archives.count()
            present = archives.filter(status=True).count()
            absent = archives.filter(status=False).count()
            percent = (present * 100.0 / total) if total > 0 else 0.0

            records = [
                {
                    'session_date': a.session_date,
                    'status': 'Present' if a.status else 'Absent',
                    'original_timestamp': a.original_timestamp,
                    'original_recorded_by': a.original_recorded_by
                }
                for a in archives
            ]

            return Response({
                'semester': int(semester),
                'subject_name': subject,
                'student_roll_number': roll,
                'student_name': archives.first().student_name if archives.exists() else None,
                'total_records': total,
                'present': present,
                'absent': absent,
                'present_percent': round(percent, 2),
                'records': records
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error("Error in ArchivalStudentDetailView: %s", str(e), exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class ArchivalAttendanceDeleteView(APIView):
    """
    Admin-only view for deleting archival attendance records by semester.
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    
    def delete(self, request):
        """
        Delete archival attendance with filters:
        - program_id: Filter by program (optional)
        - semester: Filter by semester (required)
        - start_date: Filter by session date from (optional)
        - end_date: Filter by session date to (optional)
        """
        try:
            semester = request.query_params.get('semester')
            
            if not semester:
                return Response(
                    {"error": "Semester is required for deletion"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            program_id = request.query_params.get('program_id')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            # Build query
            archives = ArchivalAttendance.objects.filter(semester=semester)
            
            if program_id:
                program = Program.objects.get(id=program_id)
                archives = archives.filter(section_name__icontains=program.name)
            
            if start_date:
                archives = archives.filter(session_date__gte=start_date)
            
            if end_date:
                archives = archives.filter(session_date__lte=end_date)
            
            count = archives.count()
            
            if count == 0:
                return Response(
                    {"error": "No archival records found matching the criteria"},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Delete the records
            archives.delete()
            
            return Response({
                "message": f"Successfully deleted {count} archival records",
                "deleted_count": count,
                "semester": semester
            }, status=status.HTTP_200_OK)
        
        except Program.DoesNotExist:
            return Response(
                {"error": "Subject not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to delete archival attendance: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ArchivalAttendanceExportView(APIView):
    """
    Admin-only view for exporting archival attendance in XLSX or CSV format.
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    
    def get(self, request):
        """
        Export archival attendance with filters:
        - program_id: Filter by program
        - semester: Filter by semester
        - start_date: Filter by session date from
        - end_date: Filter by session date to
        - format: 'xlsx' or 'csv' (default: xlsx)
        """
        try:
            # Get filters
            program_id = request.query_params.get('program_id')
            semester = request.query_params.get('semester')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            export_format = request.query_params.get('format', 'xlsx')
            
            # Build query
            archives = ArchivalAttendance.objects.all()
            
            if program_id:
                archives = archives.filter(section_name__icontains=Program.objects.get(id=program_id).name)
            
            if semester:
                archives = archives.filter(semester=semester)
            
            if start_date:
                archives = archives.filter(session_date__gte=start_date)
            
            if end_date:
                archives = archives.filter(session_date__lte=end_date)
            
            archives = archives.order_by('semester', 'section_name', 'subject_name', 'student_roll_number', 'session_date')
            
            if export_format == 'csv':
                # CSV Export
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="archival_attendance_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
                
                writer = csv.writer(response)
                writer.writerow([
                    'Roll Number', 'Student Name', 'Section', 'Subject', 
                    'Semester', 'Session Date', 'Status', 'Original Timestamp',
                    'Recorded By', 'Archived By', 'Archived At', 'Archive Note'
                ])
                
                for archive in archives:
                    archived_by = f"{archive.archived_by.first_name} {archive.archived_by.last_name}" if archive.archived_by else "Unknown"
                    writer.writerow([
                        archive.student_roll_number,
                        archive.student_name,
                        archive.section_name,
                        archive.subject_name,
                        archive.semester,
                        archive.session_date.strftime('%Y-%m-%d'),
                        'Present' if archive.status else 'Absent',
                        archive.original_timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                        archive.original_recorded_by,
                        archived_by,
                        archive.archived_at.strftime('%Y-%m-%d %H:%M:%S'),
                        archive.archive_note or ''
                    ])
                
                return response
            
            else:
                # XLSX Export using openpyxl
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    return Response(
                        {"error": "openpyxl library not installed. Please install it to export XLSX files."},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Archival Attendance"
                
                # Header styling
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                headers = [
                    'Roll Number', 'Student Name', 'Section', 'Subject', 
                    'Semester', 'Session Date', 'Status', 'Original Timestamp',
                    'Recorded By', 'Archived By', 'Archived At', 'Archive Note'
                ]
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Data rows
                for row_num, archive in enumerate(archives, 2):
                    archived_by = f"{archive.archived_by.first_name} {archive.archived_by.last_name}" if archive.archived_by else "Unknown"
                    ws.cell(row=row_num, column=1, value=archive.student_roll_number)
                    ws.cell(row=row_num, column=2, value=archive.student_name)
                    ws.cell(row=row_num, column=3, value=archive.section_name)
                    ws.cell(row=row_num, column=4, value=archive.subject_name)
                    ws.cell(row=row_num, column=5, value=archive.semester)
                    ws.cell(row=row_num, column=6, value=archive.session_date.strftime('%Y-%m-%d'))
                    ws.cell(row=row_num, column=7, value='Present' if archive.status else 'Absent')
                    ws.cell(row=row_num, column=8, value=archive.original_timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row_num, column=9, value=archive.original_recorded_by)
                    ws.cell(row=row_num, column=10, value=archived_by)
                    ws.cell(row=row_num, column=11, value=archive.archived_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row_num, column=12, value=archive.archive_note or '')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="archival_attendance_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                wb.save(response)
                
                return response
        
        except Program.DoesNotExist:
            return Response(
                {"error": "Subject not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to export archival attendance: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

